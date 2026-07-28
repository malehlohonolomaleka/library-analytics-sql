"""
build_dashboard.py
--------------------
Builds Library_Dashboard.xlsx from the query result CSVs, with a raw-data
sheet, a KPI summary using live formulas, and native Excel charts.

Run: python build_dashboard.py
Then: python /mnt/skills/public/xlsx/scripts/recalc.py Library_Dashboard.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="2E5A88")
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13, color="2E5A88")
BODY_FONT = Font(name=FONT_NAME)


def style_header(ws, row=1, cols=None):
    cols = cols or ws.max_column
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autosize(ws):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 3, 40)


def write_df(ws, df):
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT


top_books = pd.read_csv("top_books.csv")
monthly = pd.read_csv("monthly.csv")
genre = pd.read_csv("genre.csv")
overdue = pd.read_csv("overdue.csv")

wb = Workbook()

# ---- Raw data sheets (feed the KPI formulas) ----
ws_books = wb.active
ws_books.title = "Top Books"
write_df(ws_books, top_books)
style_header(ws_books)
autosize(ws_books)

bar = BarChart()
bar.title = "Top 10 Most Borrowed Books"
bar.y_axis.title = "Times Borrowed"
data_ref = Reference(ws_books, min_col=3, min_row=1, max_row=top_books.shape[0] + 1)
cats_ref = Reference(ws_books, min_col=1, min_row=2, max_row=top_books.shape[0] + 1)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.width, bar.height = 22, 12
ws_books.add_chart(bar, f"E2")

ws_monthly = wb.create_sheet("Monthly Trend")
write_df(ws_monthly, monthly)
style_header(ws_monthly)
autosize(ws_monthly)

line = LineChart()
line.title = "Monthly Loan Volume"
line.y_axis.title = "Total Loans"
data_ref = Reference(ws_monthly, min_col=2, min_row=1, max_row=monthly.shape[0] + 1)
cats_ref = Reference(ws_monthly, min_col=1, min_row=2, max_row=monthly.shape[0] + 1)
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cats_ref)
line.width, line.height = 22, 12
ws_monthly.add_chart(line, "D2")

ws_genre = wb.create_sheet("Genre Breakdown")
write_df(ws_genre, genre)
style_header(ws_genre)
autosize(ws_genre)

pie = PieChart()
pie.title = "Loans by Genre"
data_ref = Reference(ws_genre, min_col=2, min_row=1, max_row=genre.shape[0] + 1)
cats_ref = Reference(ws_genre, min_col=1, min_row=2, max_row=genre.shape[0] + 1)
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(cats_ref)
pie.width, pie.height = 14, 10
ws_genre.add_chart(pie, "E2")

ws_overdue = wb.create_sheet("Overdue Loans")
write_df(ws_overdue, overdue)
style_header(ws_overdue)
autosize(ws_overdue)

# ---- KPI Summary sheet with live formulas referencing the sheets above ----
ws_kpi = wb.create_sheet("KPI Summary", 0)
ws_kpi["A1"] = "Library Analytics — KPI Summary"
ws_kpi["A1"].font = TITLE_FONT
ws_kpi.merge_cells("A1:B1")

n_top = top_books.shape[0]
n_overdue = overdue.shape[0]
n_monthly = monthly.shape[0]

ws_kpi.append([])
ws_kpi.append(["Metric", "Value"])
style_header(ws_kpi, row=3)

kpis = [
    ("Total loans (top-10 tracked titles)", f"=SUM('Top Books'!C2:C{n_top+1})"),
    ("Most borrowed title", f"=INDEX('Top Books'!A2:A{n_top+1}, MATCH(MAX('Top Books'!C2:C{n_top+1}), 'Top Books'!C2:C{n_top+1}, 0))"),
    ("Currently overdue loans", f"=COUNTA('Overdue Loans'!A2:A{n_overdue+1})"),
    ("Longest overdue (days)", f"=MAX('Overdue Loans'!D2:D{n_overdue+1})"),
    ("Busiest month (loans)", f"=MAX('Monthly Trend'!B2:B{n_monthly+1})"),
    ("Total months tracked", f"=COUNTA('Monthly Trend'!A2:A{n_monthly+1})"),
]
for label, formula in kpis:
    ws_kpi.append([label, formula])
    row = ws_kpi.max_row
    ws_kpi.cell(row=row, column=1).font = BODY_FONT
    ws_kpi.cell(row=row, column=2).font = BODY_FONT
autosize(ws_kpi)

wb.save("Library_Dashboard.xlsx")
print("Saved Library_Dashboard.xlsx")
